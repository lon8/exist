import requests
from bs4 import BeautifulSoup
import openpyxl
import xlsxwriter
from selenium import webdriver

MINROW = 11
MAXROW = 1802

workbook = openpyxl.load_workbook('result.xlsx')

worksheet = workbook.active

wb = xlsxwriter.Workbook('test.xlsx')
ws = wb.add_worksheet()

counter = 0

def kernel(href :str, brand_xlsx :str):
    browser = webdriver.Chrome('chromedriver.exe')

    browser.get(href)

    src = browser.page_source

    soup = BeautifulSoup(src, 'lxml')

    one_more_articles = soup.find('div', id='cat-wrapper')

    if one_more_articles is not None:
        lis = one_more_articles.find_all('li')
        for li in lis:
            brand_b = li.find('b').text
            if brand_b.replace(' ', '') == brand_value.replace(' ', ''):
                link = 'https://www.exist.ru/' + li.find('a')['href']
                kernel(link, brand_xlsx)
                return

    brand = soup.find('div', class_='art').text

    print(brand)

    offers = soup.find('div', class_='allOffers').find_all('div', class_='pricerow')

    

    for offer in offers:
        avail_bool = True
        kernel(href, brand_xlsx)
        try:
            avail = offer.find('div', class_='avail').find('a', class_='gal')
        except:
            avail_bool = False
        delivery_date = offer.find('span', class_='statis').text
        price = offer.find('span', class_='price').text
        price = ''.join([x for x in price if x.isdigit()])
        print(price)
        ws.write(counter, 0, brand)
        ws.write(counter, 1, article_value ) # Артикул
        ws.write(counter, 2, avail_bool)
        ws.write(counter, 3, delivery_date)
        ws.write(counter, 4, price)
        counter += 1

    workbook.close()
    wb.close()

for it in range(MINROW, MAXROW + 1):
    article_value = worksheet.cell(row=it, column=2).value
    brand_value = worksheet.cell(row=it, column=1).value
    print(article_value)

    href = f'https://exist.ru/Price/?pcode={article_value}'
    kernel(href, brand_value)
    print('Success --> ', href)